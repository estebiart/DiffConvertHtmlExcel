import openpyxl
import re
from openpyxl.styles import PatternFill, Alignment, Font

def read_html_file(html_file):
    with open(html_file, 'r', encoding='utf-8') as entry:
        return entry.readlines()

def extract_diff_data(lines):
    data_diff = []
    initial_plus = []
    initial_minus = []
    for line in lines:
        if re.match(r'^diff', line):
            diff_part = re.sub(r'^diff --git a/(.*) b/(.*)', r'/\1', line.strip())
            if not re.search(r'\.(webp|jpg)$', diff_part):
                data_diff.append(diff_part)
                if initial_plus:
                    yield initial_plus
                    initial_plus = []
                if initial_minus:
                    yield initial_minus
                    initial_minus = []
        elif re.match(r'^\+', line):
            initial_plus.append(line.strip()[1:])
        elif re.match(r'^-', line):
            initial_minus.append(line.strip()[1:])
    if initial_plus:
        yield initial_plus
    if initial_minus:
        yield initial_minus

def create_excel_file(data_diff, data_initial_plus, data_initial_minus, excel):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["URL", "Additions", "Deletions"])
    font_diff = Font(color="FFFFFF", size=24)
    fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
    for idx, diff in enumerate(data_diff, start=2):
        cell = ws.cell(row=idx, column=1, value=diff)
        cell.fill = fill
        cell.font = font_diff
    for fila_inicio, changes in enumerate(data_initial_plus, start=2):
        text_changes = '\n'.join(changes)
        ws.cell(row=fila_inicio, column=2, value=text_changes)
    for fila_inicio, changes in enumerate(data_initial_minus, start=2):
        text_changes = '\n'.join(changes)
        ws.cell(row=fila_inicio, column=3, value=text_changes)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for fila in range(1, len(data_diff) + 1):
        ws.cell(row=fila, column=1).font = font_diff
    for row in ws.iter_rows(min_row=1, max_row=len(data_diff) + 1, max_col=3):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.column_dimensions['A'].width = 100
    ws.column_dimensions['B'].width = 100
    ws.column_dimensions['C'].width = 100
    wb.save(excel)

def process_html(html, excel):
    lines = read_html_file(html)
    data_diff = []
    data_initial_plus = []
    data_initial_minus = []
    for changes in extract_diff_data(lines):
        if len(changes) > 0:
            if len(data_diff) == len(data_initial_plus) == len(data_initial_minus):
                data_diff.append(changes[0])
            elif len(data_diff) == len(data_initial_plus) + 1:
                data_initial_plus.append(changes)
            elif len(data_diff) == len(data_initial_minus) + 1:
                data_initial_minus.append(changes)
    create_excel_file(data_diff, data_initial_plus, data_initial_minus, excel)

process_html("entry.html", "output_file.xlsx")
